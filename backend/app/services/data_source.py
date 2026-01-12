from io import BytesIO, StringIO
from typing import List
from datetime import date, datetime, timedelta
from pathlib import Path
import csv
import math

from openpyxl import load_workbook


def parse_tabular_files(files: list[tuple[str, bytes]]) -> List[dict]:
    if not files:
        raise ValueError("No data files provided.")

    tables: list[dict] = []

    for filename, content in files:
        lower = filename.lower()
        if lower.endswith(".csv"):
            table_name = _table_name_from_file(filename)
            rows = _read_csv_rows(content)
            tables.append(_build_table(table_name, rows))
        elif lower.endswith((".xlsx", ".xls")):
            tables.extend(_read_excel_tables(filename, content))
        else:
            raise ValueError(f"Unsupported file type: {filename}")

    return tables


def _normalize_rows(rows: list[dict]) -> list[dict]:
    return [{key: _normalize_value(value) for key, value in row.items()} for row in rows]


def _normalize_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, bytes):
        return value.decode(errors="ignore")
    return value


def _read_csv_rows(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    stream = StringIO(text)
    reader = csv.DictReader(stream)
    if not reader.fieldnames:
        return []
    rows = []
    for row in reader:
        rows.append({key: row.get(key) for key in reader.fieldnames})
    return rows


def _decode_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode(errors="ignore")


def _read_excel_tables(filename: str, content: bytes) -> list[dict]:
    stream = BytesIO(content)
    tables: list[dict] = []
    workbook = load_workbook(stream, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = _normalize_headers(rows[0])
        data_rows = rows[1:]
        table_rows: list[dict] = []
        for row in data_rows:
            table_rows.append(_row_to_dict(headers, row))
        table_name = f"{_table_name_from_file(filename)}::{sheet.title}"
        tables.append(_build_table(table_name, table_rows))
    return tables


def _build_table(table_name: str, rows: list[dict]) -> dict:
    if rows:
        fields = [str(col) for col in rows[0].keys()]
    else:
        fields = []
    rows = _normalize_rows(rows)
    return {
        "name": table_name,
        "fields": fields,
        "sample_rows": rows[:5],
        "rows": rows,
    }


def _table_name_from_file(filename: str) -> str:
    path = Path(filename)
    return path.stem or path.name


def _normalize_headers(headers: tuple) -> list[str]:
    normalized = []
    for index, header in enumerate(headers, start=1):
        if header is None or str(header).strip() == "":
            normalized.append(f"column_{index}")
        else:
            normalized.append(str(header))
    return normalized


def _row_to_dict(headers: list[str], row: tuple) -> dict:
    data = {}
    for index, header in enumerate(headers):
        value = row[index] if index < len(row) else None
        data[header] = value
    return data
